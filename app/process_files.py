import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def process_files(csv_file_path, excel_file_path):
    # Read both files
    csv_df = pd.read_csv(csv_file_path, dtype=str)  # Treat all columns as strings to avoid type mismatches
    excel_df = pd.read_excel(excel_file_path, dtype=str)  # Same here

    # Fill NaN values to avoid issues during comparison
    csv_df.fillna("", inplace=True)
    excel_df.fillna("", inplace=True)

    # Initialize a list to track rows with mismatches or duplicates
    mismatched_rows_excel = []
    duplicate_rows_csv = []
    duplicate_rows_excel = []

    # Compare Excel to CSV (check for rows in Excel that don't exist in CSV)
    for i, excel_row in enumerate(excel_df.values):
        if excel_row.tolist() not in csv_df.values.tolist():
            mismatched_rows_excel.append(i)
        elif excel_df.duplicated(subset=excel_df.columns).iloc[i]:  # Check for duplicates in Excel
            duplicate_rows_excel.append(i)

    # Compare CSV to Excel (check for rows in CSV that don't exist in Excel)
    for i, csv_row in enumerate(csv_df.values):
        if csv_row.tolist() not in excel_df.values.tolist():
            mismatched_rows_excel.append(i)
        elif csv_df.duplicated(subset=csv_df.columns).iloc[i]:  # Check for duplicates in CSV
            duplicate_rows_csv.append(i)

    # Save the CSV file with duplicates highlighted in red
    processed_csv_path = "processed_csv.csv"
    csv_df.to_csv(processed_csv_path, index=False)
    highlight_duplicates(processed_csv_path, duplicate_rows_csv)

    # Save the Excel file with mismatched and duplicate rows highlighted in red
    processed_excel_path = "processed_excel.xlsx"
    excel_df.to_excel(processed_excel_path, index=False)
    highlight_mismatches_and_duplicates(processed_excel_path, mismatched_rows_excel, duplicate_rows_excel)

    return processed_csv_path, processed_excel_path


def highlight_duplicates(file_path, duplicate_rows):
    # Highlight duplicate rows in red in CSV (this is a virtual process)
    # Since CSV doesn't support formatting, we simulate it for Excel.
    pass  # This step doesn't work for CSV files as CSV is plain text.


def highlight_mismatches_and_duplicates(file_path, mismatched_rows, duplicate_rows):
    # Highlight mismatched and duplicate rows in red in Excel
    wb = load_workbook(file_path)
    ws = wb.active

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Highlight mismatched rows in red
    for i in mismatched_rows:
        row_cells = ws[i + 2]  # Excel rows are 1-indexed and skip the header row
        for cell in row_cells:
            cell.fill = red_fill

    # Highlight duplicate rows in red
    for i in duplicate_rows:
        row_cells = ws[i + 2]  # Excel rows are 1-indexed and skip the header row
        for cell in row_cells:
            cell.fill = red_fill

    wb.save(file_path)
